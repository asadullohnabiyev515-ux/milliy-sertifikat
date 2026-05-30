from fastapi import FastAPI, HTTPException
from fastapi.middleware.cors import CORSMiddleware
from pydantic import BaseModel
from typing import Optional
import openpyxl
import os
from datetime import datetime

app = FastAPI(title="Milliy Sertifikat API")

# CORS — Mini App uchun ruxsat
app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_methods=["*"],
    allow_headers=["*"],
)

# Excel fayl nomi
EXCEL_FILE = "sertifikatlar.xlsx"

# ── Ma'lumotlar modeli ──────────────────────────────────────────────────────
class SertifikatRequest(BaseModel):
    ism:        str
    familiya:   str
    telefon:    str
    kurs:       str
    user_id:    Optional[int] = None

class SertifikatResponse(BaseModel):
    success:    bool
    message:    str
    sertifikat_id: Optional[str] = None

# ── Excel fayl yaratish (agar yo'q bo'lsa) ──────────────────────────────────
def excel_init():
    if not os.path.exists(EXCEL_FILE):
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Sertifikatlar"
        ws.append([
            "ID", "Ism", "Familiya", "Telefon",
            "Kurs", "Sana", "Telegram ID", "Status"
        ])
        # Ustun kengliklarini sozlash
        ws.column_dimensions["A"].width = 10
        ws.column_dimensions["B"].width = 20
        ws.column_dimensions["C"].width = 20
        ws.column_dimensions["D"].width = 15
        ws.column_dimensions["E"].width = 25
        ws.column_dimensions["F"].width = 20
        ws.column_dimensions["G"].width = 15
        ws.column_dimensions["H"].width = 15
        wb.save(EXCEL_FILE)
    return openpyxl.load_workbook(EXCEL_FILE)

# ── Keyingi ID raqamini olish ───────────────────────────────────────────────
def next_id(ws) -> int:
    rows = list(ws.iter_rows(min_row=2, values_only=True))
    if not rows or rows[-1][0] is None:
        return 1
    try:
        return int(rows[-1][0]) + 1
    except (ValueError, TypeError):
        return len(rows) + 1

# ── Endpointlar ─────────────────────────────────────────────────────────────

@app.get("/")
def root():
    return {"status": "ok", "xabar": "Milliy Sertifikat API ishlamoqda ✅"}

@app.post("/sertifikat/ariza", response_model=SertifikatResponse)
def ariza_yuborish(data: SertifikatRequest):
    """Yangi sertifikat arizasini qabul qiladi va Excel ga yozadi."""
    try:
        wb = excel_init()
        ws = wb.active
        sid = next_id(ws)
        sana = datetime.now().strftime("%Y-%m-%d %H:%M")

        ws.append([
            sid,
            data.ism,
            data.familiya,
            data.telefon,
            data.kurs,
            sana,
            data.user_id or "",
            "Kutilmoqda"
        ])
        wb.save(EXCEL_FILE)

        return SertifikatResponse(
            success=True,
            message=f"Arizangiz qabul qilindi! 🎉",
            sertifikat_id=str(sid)
        )
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@app.get("/sertifikat/tekshir/{sid}")
def sertifikat_tekshir(sid: str):
    """Sertifikat holatini tekshiradi."""
    try:
        wb = excel_init()
        ws = wb.active
        for row in ws.iter_rows(min_row=2, values_only=True):
            if str(row[0]) == sid:
                return {
                    "topildi":  True,
                    "id":       row[0],
                    "ism":      row[1],
                    "familiya": row[2],
                    "kurs":     row[4],
                    "sana":     row[5],
                    "status":   row[7],
                }
        return {"topildi": False, "xabar": "Sertifikat topilmadi"}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@app.get("/admin/royxat")
def admin_royxat():
    """Barcha arizalar ro'yxati (admin uchun)."""
    try:
        wb = excel_init()
        ws = wb.active
        natija = []
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[0] is not None:
                natija.append({
                    "id":       row[0],
                    "ism":      row[1],
                    "familiya": row[2],
                    "telefon":  row[3],
                    "kurs":     row[4],
                    "sana":     row[5],
                    "tg_id":    row[6],
                    "status":   row[7],
                })
        return {"jami": len(natija), "arizalar": natija}
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))

@app.put("/admin/status/{sid}")
def status_yangilash(sid: str, yangi_status: str):
    """Sertifikat statusini yangilaydi: Tasdiqlandi / Rad etildi."""
    try:
        wb = excel_init()
        ws = wb.active
        for row in ws.iter_rows(min_row=2):
            if str(row[0].value) == sid:
                row[7].value = yangi_status
                wb.save(EXCEL_FILE)
                return {"success": True, "xabar": f"Status '{yangi_status}' ga o'zgartirildi"}
        raise HTTPException(status_code=404, detail="Ariza topilmadi")
    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))
